"""
Excel数据读取模块
从Excel表格读取邀请名单数据
"""
import openpyxl
import xlrd
from pathlib import Path
from typing import List, Dict, Any

XLSX_MAGIC = b'PK\x03\x04'
XLS_MAGIC = b'\xd0\xcf\x11\xe0'


def _detect_format(filepath: str) -> str:
    """通过文件头魔数字节检测实际文件格式，不依赖扩展名"""
    with open(filepath, 'rb') as f:
        header = f.read(8)
    if header[:4] == XLSX_MAGIC:
        return 'xlsx'
    if header[:4] == XLS_MAGIC:
        return 'xls'
    return 'unknown'


class ExcelReader:
    """读取Excel数据"""

    def __init__(self, excel_path: str):
        self.excel_path = excel_path
        self.workbook = None
        self.worksheet = None
        self.headers = []
        self.data = []

    def read(self) -> List[Dict[str, Any]]:
        fmt = _detect_format(self.excel_path)
        print(f"[ExcelReader] 检测到文件格式: {fmt} (路径: {self.excel_path})")

        if fmt == 'xlsx':
            self._ensure_extension('.xlsx')
            self._read_xlsx()
        elif fmt == 'xls':
            self._read_xls()
        else:
            raise ValueError(
                "无法识别文件格式。该文件不是有效的 Excel 文件。"
                "支持格式: .xlsx (Office 2007+) 和 .xls (Office 97-2003)。"
                "请勿将 CSV/HTML 等文件改后缀上传。"
            )

        if not self.data:
            raise ValueError("文件中没有找到有效数据，请检查表头和数据行是否完整。")

        return self.data

    def _ensure_extension(self, ext: str) -> None:
        """确保文件有正确的扩展名，openpyxl 会检查文件名后缀"""
        if Path(self.excel_path).suffix.lower() == ext:
            return
        new_path = str(self.excel_path) + ext
        import shutil
        shutil.copy2(self.excel_path, new_path)
        print(f"[ExcelReader] 修正文件扩展名: {self.excel_path} -> {new_path}")
        self.excel_path = new_path

    def _looks_like_header(self, row_values):
        """检测一行是否像表头（短文本、列名风格）"""
        non_empty = [v for v in row_values if v is not None and str(v).strip()]
        if len(non_empty) < 2:
            return False
        for v in non_empty:
            s = str(v).strip()
            # 数据通常包含书名号、括号、数字比例高等特征
            if len(s) > 12:
                return False
            if '《' in s or '》' in s or '（' in s or '）' in s:
                return False
        return True

    def _read_xlsx(self) -> None:
        self.workbook = openpyxl.load_workbook(self.excel_path)
        self.worksheet = self.workbook.active

        header_row = None
        data_start_row = 3

        # 先尝试第1行作为表头
        row1 = None
        for row in self.worksheet.iter_rows(min_row=1, max_row=1, values_only=True):
            row1 = row
            break
        if row1 and self._looks_like_header(row1):
            header_row = row1
            data_start_row = 2
            print("[ExcelReader] 自动检测：第1行为表头")

        # 若第1行不像表头，回退到原逻辑（第2行为表头）
        if header_row is None:
            for row in self.worksheet.iter_rows(min_row=2, max_row=2, values_only=True):
                header_row = row
                break
            print("[ExcelReader] 使用第2行为表头")

        if not header_row:
            raise ValueError("无法找到表头行")

        self.headers = [str(h).strip() for h in header_row if h is not None]
        print(f"表头: {self.headers}")

        for row in self.worksheet.iter_rows(min_row=data_start_row, values_only=True):
            if not any(row):
                continue
            record = {}
            for i, header in enumerate(self.headers):
                if i < len(row):
                    record[header] = row[i]
                else:
                    record[header] = None
            self.data.append(record)

        print(f"读取了 {len(self.data)} 条数据")

    def _read_xls(self) -> None:
        self.workbook = xlrd.open_workbook(self.excel_path)
        self.worksheet = self.workbook.sheet_by_index(0)

        header_row = None
        data_start_row = 2

        # 先尝试第1行作为表头
        row1 = self.worksheet.row_values(0)
        if row1 and self._looks_like_header(row1):
            header_row = row1
            data_start_row = 1
            print("[ExcelReader] 自动检测：第1行为表头")

        # 若第1行不像表头，回退到原逻辑（第2行为表头）
        if header_row is None:
            header_row = self.worksheet.row_values(1)
            print("[ExcelReader] 使用第2行为表头")

        if not header_row:
            raise ValueError("无法找到表头行")

        self.headers = [str(h).strip() for h in header_row if h is not None and str(h).strip() != '']
        print(f"表头: {self.headers}")

        for row_idx in range(data_start_row, self.worksheet.nrows):
            row = self.worksheet.row_values(row_idx)
            if not any(row):
                continue
            record = {}
            for i, header in enumerate(self.headers):
                if i < len(row):
                    record[header] = row[i]
                else:
                    record[header] = None
            self.data.append(record)

        print(f"读取了 {len(self.data)} 条数据")

    def get_data(self) -> List[Dict[str, Any]]:
        return self.data


if __name__ == "__main__":
    reader = ExcelReader("guest_list.xlsx")
    data = reader.read()
    for i, record in enumerate(data[:3], 1):
        print(f"记录 {i}: {record}")